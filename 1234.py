"""
Scan a parent folder (and its subfolders) for .wav files.
Sum durations per folder, detect language from folder name,
and export an Excel report with per-folder and per-language summaries.

Usage:
  python wav_duration_report.py /path/to/parent/folder
  python wav_duration_report.py /path/to/parent/folder --output ~/Desktop/report.xlsx
"""

import os, sys, wave
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

LANGUAGES = ["english", "hindi", "gujarati", "marathi"]


def get_wav_duration(filepath):
    try:
        with wave.open(filepath, 'rb') as w:
            return w.getnframes() / w.getframerate()
    except Exception as e:
        print(f"  ⚠ Skipped {filepath}: {e}")
        return 0.0


def seconds_to_hms(total_seconds):
    h = int(total_seconds // 3600)
    m = int((total_seconds % 3600) // 60)
    s = total_seconds % 60
    whole_s = int(s)
    ms = int(round((s - whole_s) * 1000))
    return f"{h:02d}:{m:02d}:{whole_s:02d}.{ms:03d}"


def detect_language(folder_name):
    name_lower = folder_name.lower()
    for lang in LANGUAGES:
        if lang in name_lower:
            return lang.capitalize()
    return "Unknown"


def style_header(ws, headers, header_font, header_fill, header_align, thin_border):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def main():
    if len(sys.argv) < 2:
        print("Usage: python wav_duration_report.py /path/to/folder [--output /path/to/output.xlsx]")
        sys.exit(1)

    parent = sys.argv[1]
    if "--output" in sys.argv:
        output_path = sys.argv[sys.argv.index("--output") + 1]
    else:
        output_path = os.path.join(os.path.expanduser("~/Desktop"), "wav_duration_report.xlsx")

    if not os.path.isdir(parent):
        print(f"Error: '{parent}' is not a valid directory.")
        sys.exit(1)

    results = []

    # WAV files directly in parent folder
    root_total = 0.0
    root_count = 0
    for f in os.listdir(parent):
        if os.path.isfile(os.path.join(parent, f)) and f.lower().endswith('.wav'):
            root_total += get_wav_duration(os.path.join(parent, f))
            root_count += 1
    if root_count > 0:
        folder_name = os.path.basename(os.path.abspath(parent))
        lang = detect_language(folder_name)
        results.append((folder_name + " (root)", lang, root_count, root_total, seconds_to_hms(root_total)))

    # Subfolders
    for folder_name in sorted(os.listdir(parent)):
        folder_path = os.path.join(parent, folder_name)
        if not os.path.isdir(folder_path):
            continue
        total = 0.0
        count = 0
        for f in os.listdir(folder_path):
            if f.lower().endswith('.wav'):
                total += get_wav_duration(os.path.join(folder_path, f))
                count += 1
        if count > 0:
            lang = detect_language(folder_name)
            results.append((folder_name, lang, count, total, seconds_to_hms(total)))

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    wb = Workbook()

    # --- Sheet 1: Folder-wise Details ---
    ws1 = wb.active
    ws1.title = "Folder-wise Duration"
    headers1 = ["Folder Name", "Language", "WAV File Count", "Total Duration (hh:mm:ss.ms)"]
    style_header(ws1, headers1, header_font, header_fill, header_align, thin_border)

    for i, (name, lang, count, _, duration) in enumerate(results, 2):
        ws1.cell(row=i, column=1, value=name).border = thin_border
        ws1.cell(row=i, column=2, value=lang).border = thin_border
        ws1.cell(row=i, column=2).alignment = center
        ws1.cell(row=i, column=3, value=count).border = thin_border
        ws1.cell(row=i, column=3).alignment = center
        ws1.cell(row=i, column=4, value=duration).border = thin_border
        ws1.cell(row=i, column=4).alignment = center

    row = len(results) + 2
    grand_total_secs = sum(r[3] for r in results)
    total_vals = ["TOTAL", "", sum(r[2] for r in results), seconds_to_hms(grand_total_secs)]
    for col, val in enumerate(total_vals, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = bold_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = center

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 30

    # --- Sheet 2: Language-wise Summary ---
    ws2 = wb.create_sheet("Language-wise Summary")
    headers2 = ["Language", "Folder Count", "Total WAV Files", "Total Duration (hh:mm:ss.ms)"]
    style_header(ws2, headers2, header_font, header_fill, header_align, thin_border)

    lang_data = {}
    for _, lang, count, secs, _ in results:
        if lang not in lang_data:
            lang_data[lang] = {"folders": 0, "files": 0, "seconds": 0.0}
        lang_data[lang]["folders"] += 1
        lang_data[lang]["files"] += count
        lang_data[lang]["seconds"] += secs

    row2 = 2
    for lang in ["English", "Hindi", "Gujarati", "Marathi", "Unknown"]:
        if lang not in lang_data:
            continue
        d = lang_data[lang]
        ws2.cell(row=row2, column=1, value=lang).border = thin_border
        ws2.cell(row=row2, column=2, value=d["folders"]).border = thin_border
        ws2.cell(row=row2, column=2).alignment = center
        ws2.cell(row=row2, column=3, value=d["files"]).border = thin_border
        ws2.cell(row=row2, column=3).alignment = center
        ws2.cell(row=row2, column=4, value=seconds_to_hms(d["seconds"])).border = thin_border
        ws2.cell(row=row2, column=4).alignment = center
        row2 += 1

    total_vals2 = ["TOTAL", sum(d["folders"] for d in lang_data.values()), sum(d["files"] for d in lang_data.values()), seconds_to_hms(grand_total_secs)]
    for col, val in enumerate(total_vals2, 1):
        cell = ws2.cell(row=row2, column=col, value=val)
        cell.font = bold_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = center

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 30

    output = output_path
    wb.save(output)
    print(f"\n✅ Report saved: {output}")
    print(f"   Folders: {len(results)} | Total duration: {seconds_to_hms(grand_total_secs)}")


if __name__ == "__main__":
    main()