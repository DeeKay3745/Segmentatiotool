"""
Gujarati Timestamp Extractor
=============================
Transcribes a Gujarati WAV file using OpenAI Whisper and saves
segment-level timestamps to an Excel file.

Requirements:
    pip install openai-whisper openpyxl

Usage:
    python gujarati_timestamps.py MVI_2872.wav
    python gujarati_timestamps.py MVI_2872.wav --model medium --output timestamps.xlsx
"""

import argparse
import sys
from pathlib import Path

import whisper
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def format_time(seconds: float) -> str:
    """Convert seconds to HH:MM:SS.mmm format."""
    hrs = int(seconds // 3600)
    mins = int((seconds % 3600) // 60)
    secs = seconds % 60
    return f"{hrs:02d}:{mins:02d}:{secs:06.3f}"


def transcribe(audio_path: str, model_name: str = "medium") -> list[dict]:
    """Transcribe audio and return segments with timestamps."""
    print(f"Loading Whisper model '{model_name}' ...")
    model = whisper.load_model(model_name)

    print(f"Transcribing '{audio_path}' in Gujarati ...")
    result = model.transcribe(
        audio_path,
        language="gu",          # Gujarati language code
        task="transcribe",      # keep in original language (not translate)
        verbose=False,
    )

    segments = []
    for seg in result["segments"]:
        segments.append({
            "id": seg["id"] + 1,
            "start": seg["start"],
            "end": seg["end"],
            "text": seg["text"].strip(),
        })

    print(f"Found {len(segments)} segments.")
    return segments


def save_to_excel(segments: list[dict], output_path: str) -> None:
    """Save timestamped segments to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gujarati Timestamps"

    # ── Styles ──────────────────────────────────────────────
    header_font = Font(name="Noto Sans Gujarati", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Noto Sans Gujarati", size=11)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    time_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Headers ─────────────────────────────────────────────
    headers = ["#", "શરૂઆત (Start)", "અંત (End)", "સમયગાળો (Duration)", "ગુજરાતી ટેક્સ્ટ (Text)"]
    col_widths = [6, 18, 18, 18, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ── Data rows ───────────────────────────────────────────
    for row_idx, seg in enumerate(segments, start=2):
        duration = seg["end"] - seg["start"]
        values = [
            seg["id"],
            format_time(seg["start"]),
            format_time(seg["end"]),
            format_time(duration),
            seg["text"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = time_align if col_idx <= 4 else cell_align

    # ── Freeze header row ───────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved timestamps to '{output_path}'")


def main():
    parser = argparse.ArgumentParser(description="Extract Gujarati timestamps from a WAV file.")
    parser.add_argument("audio", help="Path to WAV audio file")
    parser.add_argument("--model", default="medium",
                        choices=["tiny", "base", "small", "medium", "large"],
                        help="Whisper model size (default: medium)")
    parser.add_argument("--output", default=None,
                        help="Output Excel path (default: <audio>_timestamps.xlsx)")
    args = parser.parse_args()

    audio_path = Path(args.audio)
    if not audio_path.exists():
        sys.exit(f"Error: File '{audio_path}' not found.")

    output_path = args.output or str(audio_path.with_suffix("")) + "_timestamps.xlsx"

    segments = transcribe(str(audio_path), args.model)
    if not segments:
        sys.exit("No speech segments detected.")

    save_to_excel(segments, output_path)
    print("Done!")


if __name__ == "__main__":
    main()